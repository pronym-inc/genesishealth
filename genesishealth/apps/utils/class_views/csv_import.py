import csv
import io

from django import forms
from django.http import HttpResponse

from openpyxl import load_workbook

from genesishealth.apps.utils.class_views import GenesisFormView
from genesishealth.apps.utils.forms import GenesisForm


EXCEL_EXTENSIONS = ('xlsx', 'xls')


class CSVImportForm(GenesisForm):
    columns = []
    line_form_class = None
    supported_file_extensions = ('xlsx', 'xls', 'csv')
    header_row_count = 1

    doc = forms.FileField()

    def clean_doc(self):
        doc = self.cleaned_data['doc']
        ext = doc.name.split('.')[-1]
        if ext not in self.supported_file_extensions:
            raise forms.ValidationError(
                'Unsupported file extension: {0}. '
                'Supported extensions are: {1}'.format(ext, self.supported_file_extensions))
        line_errors = []
        line_number = self.header_row_count + 1
        for line in self.iter_lines(doc):
            errors = self.validate_line(line)
            if errors:
                cleaned_errors = list(map(
                    lambda k: "[Line {0} - {1}] {2}".format(
                        line_number, k[0], k[1]),
                    errors.iteritems()))
                line_errors.extend(cleaned_errors)
            line_number += 1
        if line_errors:
            error_message = "<br>".join(line_errors)
            raise forms.ValidationError(
                "Errors validating lines:<br>{0}".format(error_message))
        return doc

    def get_column_names(self):
        return self.columns

    def get_extension(self, doc):
        return doc.name.split('.')[-1]

    def get_header(self, doc):
        ext = self.get_extension(doc)
        if ext == 'csv':
            return self.get_header_csv(doc)
        elif ext in EXCEL_EXTENSIONS:
            return self.get_header_xlsx(doc)
        raise Exception('Invalid extension: {0}'.format(ext))

    def get_header_csv(self, doc):
        reader = csv.reader(doc)
        output = []
        count = 0
        for line in reader:
            count += 1
            output.append(line)
            if count >= self.header_row_count:
                break
        return output

    def get_header_xlsx(self, doc):
        output = []
        wb = load_workbook(doc)
        ws = wb.active
        for row in ws.iter_rows(max_row=self.header_row_count):
            row_output = []
            for cell in row:
                if not cell:
                    break
                row_output.append(cell.value)
            output.append(row_output)
        return output

    def get_line_form_class(self):
        return self.line_form_class

    def iter_cleaned_lines(self, doc):
        for line in self.iter_lines(doc):
            form_cls = self.get_line_form_class()
            form = form_cls(line)
            if form.is_valid():
                yield form.cleaned_data

    def iter_lines(self, doc):
        extension = self.get_extension(doc)
        if extension == 'csv':
            gen = self.iter_csv_lines(doc)
        elif extension in EXCEL_EXTENSIONS:
            gen = self.iter_excel_lines(doc)
        else:
            raise Exception('Invalid extension: {0}'.format(extension))
        for i in gen:
            yield i

    def iter_csv_lines(self, doc, include_header=False):
        reader = csv.DictReader(doc, fieldnames=self.get_column_names())
        # Skip lines for header rows
        if not include_header:
            for _ in range(self.header_row_count):
                reader.next()

        for row in reader:
            yield row

    def iter_excel_lines(self, doc, include_header=False):
        wb = load_workbook(doc)
        ws = wb.active
        column_names = self.get_column_names()
        if include_header:
            start = 1
        else:
            start = self.header_row_count + 1
        rows = ws.iter_rows(
            min_row=start,
            max_col=len(column_names))
        for row in rows:
            row_output = {}
            for column_name, cell in zip(column_names, row):
                row_output[column_name] = cell.value
            if not any(row_output.values()):
                break
            yield row_output

    def validate_line(self, line):
        form_cls = self.get_line_form_class()
        form = form_cls(line)
        if not form.is_valid():
            return form.errors


class CSVImportView(GenesisFormView):
    form_class = CSVImportForm
    template_name = "utils/generic_form_csv_import.html"
    show_template_link = True

    def download_import_template(self):
        buf = io.BytesIO()
        writer = csv.writer(buf)
        for i in range(self.form_class.header_row_count - 1):
            writer.writerow([])
        writer.writerow(self.form_class.columns)
        response = HttpResponse(buf.getvalue(), content_type="text/csv")
        response['Content-Disposition'] = 'attachment; filename="template.csv"'
        return response

    def form_valid(self, form):
        for line in form.iter_cleaned_lines(form.cleaned_data['doc']):
            self.process_line(form, line)
        return super(GenesisFormView, self).form_valid(form)

    def get(self, request, *args, **kwargs):
        if request.GET.get('getTemplate'):
            return self.download_import_template()
        return super(CSVImportView, self).get(request, *args, **kwargs)

    def process_line(self, form, line):
        pass
